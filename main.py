import openpyxl

def multiple_replace(target_str, replace_values):
    for i, j in replace_values.items():
        target_str = target_str.replace(i, j)
    return target_str
def anonymity_name():
    vals = [v[0].value for v in sheet['A1:A250000']]
    i = 1
    for rec in vals:
        sheet.cell(row = i, column = 1).value = "М"
        i += 1
def anonymity_pass():
    valspass = [v[0].value for v in sheet['B1:B250000']]
    i = 1
    for rec in valspass:
        sheet.cell(row = i, column = 2).value = " "
        i += 1
def anonymity_snils():
    vals = [v[0].value for v in sheet['C1:C250000']]
    i = 1
    for rec in vals:
        a = ' '
        a = sheet.cell(row = i, column = 3).value
        replace_values = {"1": "X", "2": "X", "3": "X", "4": "X", "5": "X", "6": "X", "7": "X", "8": "X", "9": "X", "0": "X"}
        a = a[0:1] + multiple_replace(a[1:], replace_values)
        sheet.cell(row=i, column=3).value = "XXX-XXX-XXX XX"
        i += 1
def anonymity_symp():
    vals = [v[0].value for v in sheet['D1:D250000']]
    i = 1
    for rec in vals:
        a = ' '
        a = str(sheet.cell(row = i, column = 4).value)
        a = a.split(',')
        sheet.cell(row=i, column=4).value = len(a)
        if 1 < len(a) < 4:
            sheet.cell(row=i, column=4).value = 1
        elif 3 < len(a) < 7:
            sheet.cell(row=i, column=4).value = 4
        else:
            sheet.cell(row=i, column=4).value = 7
        i += 1
def anonymity_doc():
    valspass = [v[0].value for v in sheet['E1:E250000']]
    i = 1
    for rec in valspass:
        s = " "
        s = sheet.cell(row = i, column = 5).value
        if s in "Инфекционист\n Психиатр\n Психолог\n Уролог\n Гинеколог\n Венеролог\n Невролог\n Проктолог\n Логопед\n Стоматолог\n Терапевт\n Хирург\n Эндокринолог\n Фельдшер\n Онколог\n Ортопед\n Отоларинголог\n Офтальмолог\n Педиатр\n Кардиолог\n Аллерголог\n":
            sheet.cell(row=i, column=5).value = "Врач общей направленности"
        else:
            sheet.cell(row=i, column=5).value = "Врач узкой направленности"
        i += 1
def anonymity_date1():
    vals = [v[0].value for v in sheet['F1:F250000']]
    i = 1
    for rec in vals:
        a = ' '
        a = sheet.cell(row = i, column = 6).value
        if (a[5:7] == '01') or (a[5:7] == '02') or (a[5:7] == '12'):
            sheet.cell(row=i, column=6).value = 'зима'
        if (a[5:7] == '03') or (a[5:7] == '04') or (a[5:7] == '05'):
            sheet.cell(row=i, column=6).value = 'весна'
        if (a[5:7] == '06') or (a[5:7] == '07') or (a[5:7] == '08'):
            sheet.cell(row=i, column=6).value = 'лето'
        if (a[5:7] == '09') or (a[5:7] == '10') or (a[5:7] == '11'):
            sheet.cell(row=i, column=6).value = 'осень'
        i += 1
def anonymity_analys():
    vals = [v[0].value for v in sheet['G1:G250000']]
    i = 1
    for rec in vals:
        a = ' '
        a = str(sheet.cell(row = i, column = 7).value)
        a = a.split(',')
        if len(a) < 3:
            sheet.cell(row=i, column=7).value = 1
        else:
            sheet.cell(row=i, column=7).value = 5
        i += 1
def anonymity_date2():
    vals = [v[0].value for v in sheet['H1:H250000']]
    i = 1
    for rec in vals:
        a = ' '
        a = sheet.cell(row = i, column = 8).value
        if (a[0:4] == '2015') or (a[0:4] == '2016') or (a[0:4] == '2017'):
            sheet.cell(row=i, column=8).value = 2017
        else:
            sheet.cell(row=i, column=8).value = 2022
        i += 1
def anonymity_cost():
    vals = [v[0].value for v in sheet['I1:I250000']]
    i = 1
    for rec in vals:
        b = 0
        c = sheet.cell(row = i, column = 9).value
        a = int(c)
        if a < 2500:
            b = 1500
        if a >= 2500:
            b = 3750
        sheet.cell(row=i, column=9).value = b
        i += 1
def anonymity_cards():
    vals = [v[0].value for v in sheet['J1:J250000']]
    i = 1
    for rec in vals:
        a = ' '
        a = sheet.cell(row = i, column = 10).value
        replace_values = {"1": "X", "2": "X", "3": "X", "4": "X", "5": "X", "6": "X", "7": "X", "8": "X", "9": "X", "0": "X"}
        a = a[0:1] + multiple_replace(a[2:], replace_values)
        sheet.cell(row=i, column=10).value = a
        i += 1


print("Введите количество желаемых квази-идентификаторов(от 1 до 10)")
count = int(input())
print("Введите номера столбцов(от 1 до 10): 1 - ФИО; 2 - паспорт; 3 - снилс; 4 - симптомы; 5 - врач")
print("6 - дата посещения; 7 - анализы; 8 - дата выдачи анализов; 9 - стоимость анализов, 10 - номер банковской карты")
kvaz = []
for i in range(count):
    kvaz.append(int(input()))

wb = openpyxl.load_workbook(filename = 'Expenses02.xlsx')
sheet = wb['Sheet1']

anonymity_name()
anonymity_pass()
anonymity_snils()
anonymity_symp()
anonymity_doc()
anonymity_date1()
anonymity_analys()
anonymity_date2()
anonymity_cost()
anonymity_cards()

k = []
s = ' '
st = {}
flag = 0

for i in range(1, 250001):
    k.clear()
    s = ' '
    for j in range(count):
        k.append(sheet.cell(row = i, column = kvaz[j]).value)
    s = str(k)
    if s in st:
        st[s] += 1
    else:
        st[s] = 1
flag = 0
min = 250000
for i in st:
    if st[i] == 1:
        flag +=1
    if st[i] < 5:
        print(st[i], '  ',i)
    if st[i]<min:
        min = st[i]
print()
print('Уникальных строчек = ', flag)
print("k-anonymity min =", min)
for i in st:
    if st[i] == min:
        print("k-anonymity min: ",i)
wb.save('Expenses02.xlsx')
